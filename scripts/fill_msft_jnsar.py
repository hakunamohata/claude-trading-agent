"""Fill MSFT JNSAR sheet in holdings/Testing.xlsx.

Steps:
  1. Reset rows 3-6 seed values (corrupt Indian-template ~$12k seeds) to real
     MSFT H/L/C so the EMAs start at the right level and converge immediately.
  2. Append all missing trading days from the day after the sheet's last row
     through the latest cached MSFT close. Carry over the F/G/H/I formulas
     exactly (5HEMA, 5LEMA, 5CEMA, JNSAR).
  3. Compute the long-only JNSAR win rate over the full filled series:
       LONG ENTRY  when CLOSE crosses above JNSAR (next bar's open is entry)
       LONG EXIT   when CLOSE crosses below JNSAR (that day's close is exit)
       WIN if exit_price > entry_price; LOSS otherwise.
  4. Write a summary block to the right of the data (cols K-R) with all trades
     and the win-rate stats.
"""
from __future__ import annotations
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd

# Make the project importable when this is run from anywhere
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from data_fetch import fetch_many

XLSX = ROOT / "holdings" / "Testing.xlsx"
SHEET = "Msft"


def fill_and_analyze():
    print(f"Opening {XLSX}")
    wb = load_workbook(XLSX, data_only=False)
    ws = wb[SHEET]

    # ---------- Find last row with a date ----------
    last_row = None
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            last_row = r
            break
    last_date_in_sheet = ws.cell(row=last_row, column=1).value
    print(f"Last row: R{last_row} = {last_date_in_sheet} CLOSE={ws.cell(row=last_row, column=5).value}")

    # ---------- Pull MSFT OHLC ----------
    raw = fetch_many(["MSFT"], force=False)
    msft = raw["MSFT"].dropna(subset=["close"]).copy()
    print(f"MSFT cache last date: {msft.index[-1].date()} close=${msft['close'].iloc[-1]:.2f}")

    # ---------- 1. Reset seeds in rows 3-6 ----------
    # The 5-period EMA seeds (col F, G, H) at rows 3-6 are leftover Indian-template
    # values (~$12k). Replace with real MSFT H, L, C from those same dates so the
    # EMA recursion starts at the correct level.
    print("\nResetting EMA seeds in rows 3-6 to real MSFT H/L/C:")
    for r in range(3, 7):
        d = ws.cell(row=r, column=1).value
        if d is None:
            continue
        d_ts = pd.Timestamp(d).normalize()
        # Find this date in the MSFT data
        sub = msft.loc[msft.index == d_ts]
        if sub.empty:
            print(f"  R{r} {d_ts.date()}: NOT in cache (skipping)")
            continue
        H = float(sub["high"].iloc[0])
        L = float(sub["low"].iloc[0])
        C = float(sub["close"].iloc[0])
        # R3 = literal seed; R4-R6 keep their formulas which reference R3.
        if r == 3:
            ws.cell(row=r, column=6).value = H   # 5HEMA seed = first HIGH
            ws.cell(row=r, column=7).value = L   # 5LEMA seed = first LOW
            ws.cell(row=r, column=8).value = C   # 5CEMA seed = first CLOSE
            print(f"  R{r} {d_ts.date()}: SET seeds H={H:.2f} L={L:.2f} C={C:.2f}")
        else:
            # Leave the formulas as-is; they'll recompute from R3 seeds.
            print(f"  R{r} {d_ts.date()}: formula preserved")

    # ---------- 2. Append missing rows ----------
    last_date_ts = pd.Timestamp(last_date_in_sheet).normalize()
    new_data = msft.loc[msft.index > last_date_ts]
    print(f"\nAppending {len(new_data)} new trading days:")
    print(f"  From {new_data.index[0].date()}  to {new_data.index[-1].date()}")

    row = last_row + 1
    for idx, ohlc in new_data.iterrows():
        ws.cell(row=row, column=1).value = idx.to_pydatetime()
        ws.cell(row=row, column=2).value = float(ohlc["open"])
        ws.cell(row=row, column=3).value = float(ohlc["high"])
        ws.cell(row=row, column=4).value = float(ohlc["low"])
        ws.cell(row=row, column=5).value = float(ohlc["close"])
        # Formulas — mirror the rows above. Reference (row-1) for prior EMA.
        ws.cell(row=row, column=6).value = f"=F{row-1}+(C{row}-F{row-1})*2/6"   # 5HEMA
        ws.cell(row=row, column=7).value = f"=G{row-1}+(D{row}-G{row-1})*2/6"   # 5LEMA
        ws.cell(row=row, column=8).value = f"=H{row-1}+(E{row}-H{row-1})*2/6"   # 5CEMA
        # JNSAR mirrors the existing pattern: AVERAGE over a 5x3 rectangle of EMAs
        ws.cell(row=row, column=9).value = f"=AVERAGE(F{row-4}:H{row})"
        row += 1

    new_last_row = row - 1
    print(f"  Sheet now has data through row {new_last_row}")

    # ---------- 3. Compute JNSAR signals + win rate (in Python) ----------
    # We can't read the formula-derived values back without saving + reopening, so
    # we recompute the EMAs and JNSAR locally from the OHLC series.
    print("\nComputing JNSAR signals and win rate (locally)...")

    # Build the full OHLC series from R3 onward (using real MSFT data)
    dates = []
    O, H, L, C = [], [], [], []
    for r in range(3, new_last_row + 1):
        d = ws.cell(row=r, column=1).value
        if d is None:
            continue
        ts = pd.Timestamp(d).normalize()
        sub = msft.loc[msft.index == ts]
        if sub.empty:
            continue
        dates.append(ts)
        O.append(float(sub["open"].iloc[0]))
        H.append(float(sub["high"].iloc[0]))
        L.append(float(sub["low"].iloc[0]))
        C.append(float(sub["close"].iloc[0]))

    df = pd.DataFrame({"open": O, "high": H, "low": L, "close": C}, index=pd.DatetimeIndex(dates))

    # 5-period EMA with alpha = 2/6 = 0.3333... matching the sheet's formula
    alpha = 2.0 / 6.0

    def ema_with_seed(values, seed, alpha):
        out = [seed]
        for v in values[1:]:
            out.append(out[-1] + (v - out[-1]) * alpha)
        return out

    df["hema"] = ema_with_seed(df["high"].tolist(), df["high"].iloc[0], alpha)
    df["lema"] = ema_with_seed(df["low"].tolist(),  df["low"].iloc[0],  alpha)
    df["cema"] = ema_with_seed(df["close"].tolist(), df["close"].iloc[0], alpha)

    # JNSAR = AVERAGE of last 5 rows' HEMA, LEMA, CEMA (15 values).
    # For rows where we have fewer than 5 prior rows, use what's available.
    jnsar = []
    for i in range(len(df)):
        lo = max(0, i - 4)
        block = df.iloc[lo:i+1][["hema", "lema", "cema"]].values.flatten()
        jnsar.append(block.mean())
    df["jnsar"] = jnsar

    # Long-only signal logic:
    #   Entry: CLOSE crosses ABOVE JNSAR (today close > jnsar AND yesterday close <= yesterday jnsar)
    #   Exit:  CLOSE crosses BELOW JNSAR (today close < jnsar AND yesterday close >= yesterday jnsar)
    # Entry price = next day's OPEN (realistic execution); if no next day, skip.
    # Exit price  = the exit-day CLOSE.
    df["pos_prev"] = (df["close"].shift(1) > df["jnsar"].shift(1))
    df["pos_now"]  = (df["close"] > df["jnsar"])
    df["entry"] = (df["pos_now"]) & (~df["pos_prev"].fillna(False))
    df["exit"]  = (~df["pos_now"]) & (df["pos_prev"].fillna(False))

    trades = []
    in_pos = False
    entry_date = entry_price = None
    df_arr = df.reset_index().rename(columns={"index": "date"}).to_dict("records")
    for i, row in enumerate(df_arr):
        if not in_pos and row["entry"]:
            # Entry on NEXT bar's open
            if i + 1 < len(df_arr):
                entry_date = df_arr[i+1]["date"]
                entry_price = df_arr[i+1]["open"]
                in_pos = True
        elif in_pos and row["exit"]:
            exit_date = row["date"]
            exit_price = row["close"]
            ret_pct = (exit_price / entry_price - 1) * 100
            trades.append({
                "entry_date": entry_date.date(),
                "entry_price": round(entry_price, 2),
                "exit_date": exit_date.date(),
                "exit_price": round(exit_price, 2),
                "return_pct": round(ret_pct, 2),
                "win": ret_pct > 0,
            })
            in_pos = False
    # Open position at series end — close at last close for reporting
    if in_pos:
        exit_date = df_arr[-1]["date"]
        exit_price = df_arr[-1]["close"]
        ret_pct = (exit_price / entry_price - 1) * 100
        trades.append({
            "entry_date": entry_date.date(),
            "entry_price": round(entry_price, 2),
            "exit_date": str(exit_date.date()) + " (OPEN)",
            "exit_price": round(exit_price, 2),
            "return_pct": round(ret_pct, 2),
            "win": ret_pct > 0,
        })

    n_trades = len(trades)
    if n_trades == 0:
        print("No JNSAR trades detected.")
    else:
        wins = sum(1 for t in trades if t["win"])
        win_rate = wins / n_trades * 100
        avg_ret = sum(t["return_pct"] for t in trades) / n_trades
        avg_win = sum(t["return_pct"] for t in trades if t["win"]) / wins if wins else 0
        losses = [t for t in trades if not t["win"]]
        avg_loss = sum(t["return_pct"] for t in losses) / len(losses) if losses else 0
        sum_ret = sum(t["return_pct"] for t in trades)
        print(f"  Trades: {n_trades}  Wins: {wins}  Losses: {n_trades-wins}")
        print(f"  Win rate: {win_rate:.1f}%")
        print(f"  Avg return: {avg_ret:+.2f}%   Avg win: {avg_win:+.2f}%   Avg loss: {avg_loss:+.2f}%")
        print(f"  Total return (sum of trades): {sum_ret:+.2f}%")

    # ---------- 4. Write summary block to the sheet (cols K-Q starting R2) ----------
    summary_col = 11   # column K
    ws.cell(row=2, column=summary_col).value     = "JNSAR LONG-ONLY BACKTEST"
    ws.cell(row=2, column=summary_col+1).value   = "Computed by fill_msft_jnsar.py"
    ws.cell(row=3, column=summary_col).value     = "#"
    ws.cell(row=3, column=summary_col+1).value   = "Entry date"
    ws.cell(row=3, column=summary_col+2).value   = "Entry"
    ws.cell(row=3, column=summary_col+3).value   = "Exit date"
    ws.cell(row=3, column=summary_col+4).value   = "Exit"
    ws.cell(row=3, column=summary_col+5).value   = "Return %"
    ws.cell(row=3, column=summary_col+6).value   = "W/L"
    for i, t in enumerate(trades, 1):
        rr = 3 + i
        ws.cell(row=rr, column=summary_col).value   = i
        ws.cell(row=rr, column=summary_col+1).value = str(t["entry_date"])
        ws.cell(row=rr, column=summary_col+2).value = t["entry_price"]
        ws.cell(row=rr, column=summary_col+3).value = str(t["exit_date"])
        ws.cell(row=rr, column=summary_col+4).value = t["exit_price"]
        ws.cell(row=rr, column=summary_col+5).value = t["return_pct"]
        ws.cell(row=rr, column=summary_col+6).value = "WIN" if t["win"] else "LOSS"

    # Stats block below trades
    stats_row = 4 + len(trades) + 2
    if n_trades > 0:
        wins = sum(1 for t in trades if t["win"])
        win_rate = wins / n_trades * 100
        avg_ret = sum(t["return_pct"] for t in trades) / n_trades
        losses = [t for t in trades if not t["win"]]
        avg_win  = sum(t["return_pct"] for t in trades if t["win"]) / wins if wins else 0
        avg_loss = sum(t["return_pct"] for t in losses) / len(losses) if losses else 0
        sum_ret  = sum(t["return_pct"] for t in trades)
        rows = [
            ("Total trades",       n_trades),
            ("Wins",               wins),
            ("Losses",             n_trades - wins),
            ("Win rate %",         round(win_rate, 1)),
            ("Avg return per trade %", round(avg_ret, 2)),
            ("Avg win %",          round(avg_win, 2)),
            ("Avg loss %",         round(avg_loss, 2)),
            ("Sum of returns %",   round(sum_ret, 2)),
            ("Buy-and-hold % (entry-to-end)",
                round((df["close"].iloc[-1] / df["close"].iloc[0] - 1) * 100, 2)),
        ]
        for i, (k, v) in enumerate(rows):
            ws.cell(row=stats_row + i, column=summary_col).value     = k
            ws.cell(row=stats_row + i, column=summary_col + 1).value = v

    # ---------- Save ----------
    wb.save(XLSX)
    print(f"\nSaved {XLSX}")
    print(f"Summary block written starting at K2.")
    return trades, df


if __name__ == "__main__":
    fill_and_analyze()
